#!/usr/bin/env python3
"""
Parst Gebäudeplan-PDFs via Kachel-OCR und erzeugt gebaeudedaten.xlsx
Erkennt Raumnummern, Flächen (HNF), Bezeichnungen und Barcodes.

Nutzung:
  python parse_gebaeudeplan.py                        # alle *.pdf aus plaene/
  python parse_gebaeudeplan.py plan1.pdf plan2.pdf    # bestimmte PDFs
  python parse_gebaeudeplan.py -o output.xlsx *.pdf   # Output-Datei angeben
  python parse_gebaeudeplan.py --cpu 80               # CPU auf 80% begrenzen (default: 90%)
  python parse_gebaeudeplan.py --append                # bestehende Sheets in xlsx beibehalten
"""

import re
import sys
import os
import glob
import fitz  # PyMuPDF
import easyocr
from PIL import Image
import openpyxl

OUTPUT_FILE = "gebaeudedaten.xlsx"
DPI = 600
TILE_SIZE = 1500
TILE_OVERLAP = 200

# --- Pattern ---
# Raumnummern: X.YY oder X.YYa — Ganzzahlteil wird pro PDF aus Metadaten gefiltert
ROOM_PATTERN = re.compile(r'^(\d{1,2}\.\d{2,3}[a-z]?)$')
SPECIAL_ROOM_PATTERN = re.compile(r'^(TH\s*[A-Z]|A\s*\d+)$')
BARCODE_PATTERN = re.compile(r'[Bb]arcode[:\s;]+(\S+)')
AREA_PATTERN = re.compile(r'(\d{1,3}[\.,]\d{1,2})\s*m[2zZ²]?')

# Bekannte Maße (Fenster/Türen), die NICHT als Raumnummern erkannt werden sollen
KNOWN_DIMENSIONS = {'0.56', '0.68', '0.72', '0.81', '0.83', '0.88', '0.90', '0.91',
                    '0.93', '0.94', '0.98', '1.16', '1.18', '1.34', '1.90',
                    '1.92', '1.94', '1.95', '1.96', '2.08', '2.09', '2.12',
                    '2.14', '2.18', '2.20', '2.22', '108.12', '38.20',
                    '46.59', '16.04', '16.00'}

# Bekannte Nutzungen + OCR-Varianten
NUTZUNG_MAP = {
    'büro': 'Büro', 'buro': 'Büro', 'biro': 'Büro', 'buiro': 'Büro',
    'lager': 'Lager', 'lzger': 'Lager', 'lacei': 'Lager', 'laqer': 'Lager',
    'flur': 'Flur',
    'sanitärraum': 'Sanitärraum', 'sanitarraum': 'Sanitärraum',
    'teeküche': 'Teeküche', 'teekuche': 'Teeküche', 'teeküche': 'Teeküche',
    'treppenhaus': 'Treppenhaus', 'trecjenals': 'Treppenhaus',
    'aufzug': 'Aufzug',
    'druckerei': 'Druckerei', 'daickere': 'Druckerei', 'druckere': 'Druckerei',
    'besprechung': 'Besprechung', 'besprechunq': 'Besprechung',
    'fitnessraum': 'Fitnessraum', 'fitnesraum': 'Fitnessraum',
    'umkleideraum': 'Umkleideraum',
    'kopierer': 'Kopierer', 'kopieier': 'Kopierer',
    'wegefläche': 'Wegefläche', 'wegeflache': 'Wegefläche',
    'werkstatt': 'Werkstatt',
    'lager/werkstatt': 'Lager/Werkstatt',
    'serverraum': 'Serverraum',
    'technik': 'Technik',
    'archiv': 'Archiv',
    'küche': 'Küche', 'kuche': 'Küche',
}


def match_nutzung(text):
    """Versucht, einen OCR-Text einer bekannten Nutzung zuzuordnen."""
    t = text.strip().lower()
    if t in NUTZUNG_MAP:
        return NUTZUNG_MAP[t]
    for key, val in NUTZUNG_MAP.items():
        if key in t or t in key:
            return val
    return None


def distance(x1, y1, x2, y2):
    return ((x1 - x2) ** 2 + (y1 - y2) ** 2) ** 0.5


def weighted_distance(x1, y1, x2, y2, dx_weight=3):
    """Gewichtete Distanz: horizontaler Abstand zählt stärker (Daten stehen untereinander)."""
    dx = abs(x1 - x2) * dx_weight
    dy = y1 - y2
    return (dx ** 2 + dy ** 2) ** 0.5


def count_tiles_for_image(W, H):
    """Berechnet die Anzahl Kacheln für ein Bild der Größe WxH."""
    import math
    step = TILE_SIZE - TILE_OVERLAP
    return math.ceil(W / step) * math.ceil(H / step)


# Globaler Fortschritt über alle PDFs
_global_progress = {'done': 0, 'total': 0, 't_start': 0}


def ocr_tiled(img_path, reader, room_prefix=None):
    """OCR auf Kacheln für bessere Erkennung bei großen Bildern."""
    import time
    img = Image.open(img_path)
    W, H = img.size
    local_total = count_tiles_for_image(W, H)
    print(f"  Bild: {W}x{H}, {local_total} Kacheln à {TILE_SIZE}px")

    gp = _global_progress
    all_results = []
    # Live-Zuordnung: Räume mit Position + zugeordnete Daten
    live_rooms = {}      # {nr: {x, y, nutzung, flaeche, barcode}}
    live_areas = []      # [{area, x, y}]
    live_nutzungen = []  # [{nutzung, x, y}]
    live_barcodes = []   # [{code, x, y}]
    live_hnf = []        # [{x, y}]
    LIVE_MAX_DIST = 350
    live_printed = {}    # {nr: "last_printed_line"} — nur bei Änderung neu ausgeben
    used_barcodes = {}   # {barcode: raumnr} — Duplikat-Erkennung
    used_flaechen = {}   # {flaeche: raumnr} — Duplikat-Warnung
    live_warnings = []   # gesammelte Warnungen

    def _live_find_nearest_room(x, y, kind=None):
        """Findet den nächsten Raum. kind='flaeche'|'nutzung'|'barcode' für Reihenfolge-Check."""
        best, best_d = None, LIVE_MAX_DIST
        for nr, r in live_rooms.items():
            dy = y - r['y']
            # Daten liegen immer UNTER der Raumnummer
            if dy < 0:
                continue
            # Reihenfolge: Raumnr > Fläche > Nutzung > Barcode
            # Barcode muss weiter unten sein als Fläche, etc.
            if kind == 'flaeche' and dy > 250:
                continue  # Fläche direkt unter Raumnr
            if kind == 'nutzung' and dy > 350:
                continue
            if kind == 'barcode' and dy > 450:
                continue
            d = weighted_distance(x, y, r['x'], r['y'])
            if d < best_d:
                best_d = d
                best = nr
        return best

    def _live_try_assign_new(x, y, kind, value):
        """Versucht ein neues Area/Nutzung/Barcode dem nächsten Raum zuzuordnen."""
        nr = _live_find_nearest_room(x, y, kind=kind)
        if nr and not live_rooms[nr][kind]:
            # Duplikat-Prüfung
            if kind == 'barcode':
                if value in used_barcodes:
                    warn = f"FEHLER: Barcode '{value}' doppelt! Raum {used_barcodes[value]} und {nr}"
                    live_warnings.append(warn)
                    print(f"\n    *** {warn}", end='', flush=True)
                    return  # Nicht zuordnen
                used_barcodes[value] = nr
            elif kind == 'flaeche':
                if value in used_flaechen:
                    warn = f"Hinweis: Fläche {value} m² identisch mit Raum {used_flaechen[value]}"
                    live_warnings.append(warn)
                    print(f"\n    ?   {warn}", end='', flush=True)
                used_flaechen[value] = nr
            live_rooms[nr][kind] = value
            _live_print_room(nr)

    def _live_print_room(nr):
        r = live_rooms[nr]
        parts = [f"Raum {nr:<10}"]
        if r['flaeche']:  parts.append(f"{r['flaeche']:>8} m²")
        if r['nutzung']:  parts.append(r['nutzung'])
        if r['barcode']:  parts.append(f"BC:{r['barcode']}")
        line = '  '.join(parts)
        if live_printed.get(nr) != line:
            live_printed[nr] = line
            print(f"\n    + {line}", end='', flush=True)

    ty = 0
    tile_count = 0
    t_start = time.time()
    while ty < H:
        tx = 0
        while tx < W:
            tile_count += 1
            gp['done'] += 1
            # Gesamt-Fortschritt
            g_pct = gp['done'] * 100 // gp['total'] if gp['total'] else 0
            if gp['done'] > 1:
                g_elapsed = time.time() - gp['t_start']
                g_avg = g_elapsed / (gp['done'] - 1)
                g_remaining = g_avg * (gp['total'] - gp['done'])
                g_eta = f"~{_format_duration(g_remaining)}"
            else:
                g_eta = "..."
            rooms_info = f" | Räume: {len(live_rooms)}" if live_rooms else ""
            print(f"\r  Kachel {tile_count}/{local_total} | Gesamt: {gp['done']}/{gp['total']} ({g_pct}%) — Rest: {g_eta}{rooms_info}   ", end='', flush=True)
            x2 = min(tx + TILE_SIZE, W)
            y2 = min(ty + TILE_SIZE, H)
            crop = img.crop((tx, ty, x2, y2))
            crop_path = '_temp_tile.png'
            crop.save(crop_path)
            results = reader.readtext(crop_path, paragraph=False)
            for bbox, text, conf in results:
                gx = bbox[0][0] + tx
                gy = bbox[0][1] + ty
                t = text.strip()
                all_results.append({'x': gx, 'y': gy, 'text': t, 'conf': conf})

                # Live-Klassifizierung
                # Raumnummer?
                if ROOM_PATTERN.match(t) and conf > 0.4 and t not in KNOWN_DIMENSIONS:
                    if room_prefix is None or t.split('.')[0] == str(room_prefix):
                        if t not in live_rooms:
                            live_rooms[t] = {'x': gx, 'y': gy, 'nutzung': '', 'flaeche': '', 'barcode': ''}
                            # Rückwärts: bereits bekannte Daten zuordnen
                            # Daten müssen unterhalb der Raumnr liegen (a['y'] > gy)
                            for a in live_areas:
                                a_dy = a['y'] - gy
                                if not live_rooms[t]['flaeche'] and 0 < a_dy < 250 and weighted_distance(gx, gy, a['x'], a['y']) < LIVE_MAX_DIST:
                                    val = a['area']
                                    if val in used_flaechen:
                                        warn = f"Hinweis: Fläche {val} m² identisch mit Raum {used_flaechen[val]}"
                                        live_warnings.append(warn)
                                        print(f"\n    ?   {warn}", end='', flush=True)
                                    used_flaechen.setdefault(val, t)
                                    live_rooms[t]['flaeche'] = val
                            for n in live_nutzungen:
                                n_dy = n['y'] - gy
                                if not live_rooms[t]['nutzung'] and 0 < n_dy < 350 and weighted_distance(gx, gy, n['x'], n['y']) < 300:
                                    live_rooms[t]['nutzung'] = n['nutzung']
                            for b in live_barcodes:
                                b_dy = b['y'] - gy
                                if not live_rooms[t]['barcode'] and 0 < b_dy < 450 and weighted_distance(gx, gy, b['x'], b['y']) < LIVE_MAX_DIST:
                                    bc = b['code']
                                    if bc in used_barcodes:
                                        warn = f"FEHLER: Barcode '{bc}' doppelt! Raum {used_barcodes[bc]} und {t}"
                                        live_warnings.append(warn)
                                        print(f"\n    *** {warn}", end='', flush=True)
                                    else:
                                        used_barcodes[bc] = t
                                        live_rooms[t]['barcode'] = bc
                            _live_print_room(t)
                    continue

                # HNF-Marker?
                if t in ('HNF', 'HNF:', 'HNF(1,5m)'):
                    live_hnf.append({'x': gx, 'y': gy})
                    continue

                # Fläche?
                am = AREA_PATTERN.match(t)
                if am and conf > 0.5:
                    val = am.group(1).replace(',', '.')
                    fval = float(val)
                    if 3 < fval < 300:
                        live_areas.append({'area': val, 'x': gx, 'y': gy})
                        if any(distance(gx, gy, h['x'], h['y']) < 150 for h in live_hnf):
                            _live_try_assign_new(gx, gy, 'flaeche', val)
                        continue

                # Reiner Zahlenwert neben HNF?
                plain_num = re.match(r'^(\d{2,3}\.\d{2})$', t)
                if plain_num and conf > 0.6 and t not in KNOWN_DIMENSIONS:
                    val = plain_num.group(1)
                    fval = float(val)
                    if 3 < fval < 300 and any(distance(gx, gy, h['x'], h['y']) < 120 for h in live_hnf):
                        live_areas.append({'area': val, 'x': gx, 'y': gy})
                        _live_try_assign_new(gx, gy, 'flaeche', val)
                        continue

                # Barcode?
                bm = BARCODE_PATTERN.search(t)
                if bm and conf > 0.4:
                    code = bm.group(1).rstrip(';.,')
                    live_barcodes.append({'code': code, 'x': gx, 'y': gy})
                    _live_try_assign_new(gx, gy, 'barcode', code)
                    continue

                # Nutzung?
                nutz = match_nutzung(t)
                if nutz and conf > 0.4:
                    live_nutzungen.append({'nutzung': nutz, 'x': gx, 'y': gy})
                    _live_try_assign_new(gx, gy, 'nutzung', nutz)

            tx += TILE_SIZE - TILE_OVERLAP
        ty += TILE_SIZE - TILE_OVERLAP

    elapsed_total = time.time() - t_start
    print(f"\r  {tile_count} Kacheln, {len(all_results)} Textblöcke, {len(live_rooms)} Räume erkannt ({_format_duration(elapsed_total)})          ")

    # Warnungen zusammenfassen
    if live_warnings:
        print(f"\n  {'!'*50}")
        print(f"  {len(live_warnings)} Warnung(en):")
        for w in live_warnings:
            print(f"    - {w}")
        print(f"  {'!'*50}")

    # Deduplizieren
    unique = []
    for r in all_results:
        dup = False
        for i, u in enumerate(unique):
            if abs(r['x'] - u['x']) < 40 and abs(r['y'] - u['y']) < 40:
                if r['text'] == u['text']:
                    if r['conf'] > u['conf']:
                        unique[i] = r
                    dup = True
                    break
        if not dup:
            unique.append(r)

    print(f"  Nach Deduplizierung: {len(unique)} eindeutige Textblöcke")
    return unique


def classify_texts(ocr_results, room_prefix=None):
    """Klassifiziert OCR-Ergebnisse in Raumnummern, Barcodes, Flächen, Nutzungen."""
    rooms = []
    barcodes = []
    areas = []
    nutzungen = []
    hnf_markers = []

    for r in ocr_results:
        text = r['text']
        x, y = r['x'], r['y']
        conf = r['conf']

        # Raumnummer
        if ROOM_PATTERN.match(text) and conf > 0.4 and text not in KNOWN_DIMENSIONS:
            # Wenn Gebäude-Prefix bekannt, nur passende Raumnummern akzeptieren
            if room_prefix is not None:
                nr_prefix = text.split('.')[0]
                if nr_prefix != str(room_prefix):
                    continue
            rooms.append({'nr': text, 'x': x, 'y': y, 'conf': conf})
            continue

        if SPECIAL_ROOM_PATTERN.match(text) and conf > 0.5:
            rooms.append({'nr': text, 'x': x, 'y': y, 'conf': conf})
            continue

        # Barcode
        m = BARCODE_PATTERN.search(text)
        if m and conf > 0.4:
            code = m.group(1).rstrip(';.,')
            barcodes.append({'code': code, 'x': x, 'y': y})
            continue

        # HNF-Marker
        if text in ('HNF', 'HNF:', 'HNF(1,5m)'):
            hnf_markers.append({'x': x, 'y': y})
            continue

        # Fläche mit m²-Suffix
        am = AREA_PATTERN.match(text)
        if am and conf > 0.5:
            val = am.group(1).replace(',', '.')
            fval = float(val)
            if 3 < fval < 300:
                areas.append({'area': val, 'x': x, 'y': y})
                continue

        # Reiner Zahlenwert neben HNF
        plain_num = re.match(r'^(\d{2,3}\.\d{2})$', text)
        if plain_num and conf > 0.6 and text not in KNOWN_DIMENSIONS:
            val = plain_num.group(1)
            fval = float(val)
            if 3 < fval < 300:
                near_hnf = any(distance(x, y, h['x'], h['y']) < 120 for h in hnf_markers)
                if near_hnf:
                    areas.append({'area': val, 'x': x, 'y': y})
                    continue

        # Nutzung
        nutz = match_nutzung(text)
        if nutz and conf > 0.4:
            nutzungen.append({'nutzung': nutz, 'x': x, 'y': y})

    print(f"  Klassifiziert: {len(rooms)} Räume, {len(barcodes)} Barcodes, "
          f"{len(areas)} Flächen, {len(nutzungen)} Nutzungen, {len(hnf_markers)} HNF-Marker")
    return rooms, barcodes, areas, nutzungen, hnf_markers


def assign_to_rooms(rooms, barcodes, areas, nutzungen, hnf_markers):
    """Ordnet Barcodes, Flächen und Nutzungen den nächsten Raumnummern zu."""
    MAX_DIST_BARCODE = 350
    MAX_DIST_AREA = 350
    MAX_DIST_NUTZUNG = 300

    result = {}
    for room in rooms:
        nr = room['nr']
        if nr in result:
            if room['conf'] > result[nr]['conf']:
                result[nr] = {
                    'raumnr': nr, 'flaeche': '', 'nutzung': '', 'barcode': '',
                    'x': room['x'], 'y': room['y'], 'conf': room['conf']
                }
        else:
            result[nr] = {
                'raumnr': nr, 'flaeche': '', 'nutzung': '', 'barcode': '',
                'x': room['x'], 'y': room['y'], 'conf': room['conf']
            }

    room_list = list(result.values())

    def find_nearest_room(x, y, max_dist, max_dy=None):
        best = None
        best_dist = max_dist
        for r in room_list:
            dy = y - r['y']
            # Daten liegen immer UNTER der Raumnummer
            if dy < 0:
                continue
            if max_dy and dy > max_dy:
                continue
            d = weighted_distance(x, y, r['x'], r['y'])
            if d < best_dist:
                best_dist = d
                best = r
        return best

    # Reihenfolge v.o.n.u.: Raumnr > Fläche > Nutzung > Barcode
    for area in areas:
        has_hnf = any(distance(area['x'], area['y'], h['x'], h['y']) < 150 for h in hnf_markers)
        if not has_hnf:
            continue
        room = find_nearest_room(area['x'], area['y'], MAX_DIST_AREA, max_dy=250)
        if room and not room['flaeche']:
            room['flaeche'] = area['area']

    for nutz in nutzungen:
        room = find_nearest_room(nutz['x'], nutz['y'], MAX_DIST_NUTZUNG, max_dy=350)
        if room and not room['nutzung']:
            room['nutzung'] = nutz['nutzung']

    for bc in barcodes:
        room = find_nearest_room(bc['x'], bc['y'], MAX_DIST_BARCODE, max_dy=450)
        if room and not room['barcode']:
            room['barcode'] = bc['code']

    return result


def extract_metadata(pdf_file):
    """Extrahiert Gebäude und Etage aus PDF-Text oder Dateiname."""
    gebaeude = None
    etage = None
    geb_nr = None

    # 1. Versuch: aus dem PDF-Text (Helvetica-Font)
    doc = fitz.open(pdf_file)
    page = doc[0]
    blocks = page.get_text('dict')['blocks']
    for b in blocks:
        if 'lines' not in b:
            continue
        for line in b['lines']:
            for span in line['spans']:
                if span['font'] == 'Helvetica':
                    text = span['text'].strip()
                    m = re.search(r'Geb.ude\s+(\d+),\s+(\S+\s*\d*)', text)
                    if m:
                        geb_nr = int(m.group(1))
                        gebaeude = f'Gebäude {geb_nr}'
                        etage = m.group(2).strip()
    doc.close()

    # 2. Fallback: aus dem Dateinamen
    # z.B. "2018 - Gebaeude 1 - 1.Obergeschoss.pdf"
    basename = os.path.basename(pdf_file)
    if not gebaeude:
        m = re.search(r'Geb.ude\s*(\d+)', basename, re.IGNORECASE)
        if m:
            geb_nr = int(m.group(1))
            gebaeude = f'Gebäude {geb_nr}'
    if not etage:
        # Etage aus Dateiname: "1.Obergeschoss" -> "OG 1", "Erdgeschoss" -> "EG",
        # "1.Untergeschoss" -> "UG 1"
        m = re.search(r'(\d+)\.\s*Obergeschoss', basename, re.IGNORECASE)
        if m:
            etage = f'OG {m.group(1)}'
        elif re.search(r'Erdgeschoss', basename, re.IGNORECASE):
            etage = 'EG'
        m = re.search(r'(\d+)\.\s*Untergeschoss', basename, re.IGNORECASE)
        if m:
            etage = f'UG {m.group(1)}'

    return gebaeude, etage, geb_nr


def process_single_pdf(pdf_file, reader):
    """Verarbeitet eine einzelne PDF und gibt (gebaeude, etage, rooms) zurück."""
    print(f"\n{'='*60}")
    print(f"Verarbeite: {pdf_file}")
    print(f"{'='*60}")

    if not os.path.exists(pdf_file):
        print(f"  FEHLER: Datei nicht gefunden!")
        return None, None, {}

    # Metadaten
    gebaeude, etage, geb_nr = extract_metadata(pdf_file)
    print(f"  Gebäude: {gebaeude}, Etage: {etage}")

    # PDF rendern
    print(f"  Rendere bei {DPI} DPI...")
    doc = fitz.open(pdf_file)
    page = doc[0]
    pix = page.get_pixmap(dpi=DPI)
    img_path = "_temp_plan_ocr.png"
    pix.save(img_path)

    # OCR
    print(f"  Starte Kachel-OCR...")
    ocr_results = ocr_tiled(img_path, reader, room_prefix=geb_nr)

    # Klassifizieren (mit Gebäude-Prefix-Filter)
    rooms, barcodes, areas, nutzungen, hnf_markers = classify_texts(ocr_results, room_prefix=geb_nr)

    # Zuordnen
    room_data = assign_to_rooms(rooms, barcodes, areas, nutzungen, hnf_markers)

    # Ergebnis anzeigen
    sorted_rooms = sorted(room_data.values(), key=sort_key)
    print(f"\n  {'Raum':<10} {'Fläche':>10} {'Nutzung':<20} {'Barcode':<10}")
    print(f"  {'-'*55}")
    for r in sorted_rooms:
        print(f"  {r['raumnr']:<10} {r['flaeche']:>10} {r['nutzung']:<20} {r['barcode']:<10}")

    with_barcode = sum(1 for r in sorted_rooms if r['barcode'])
    with_area = sum(1 for r in sorted_rooms if r['flaeche'])
    with_nutzung = sum(1 for r in sorted_rooms if r['nutzung'])
    print(f"\n  Statistik: {len(sorted_rooms)} Räume, "
          f"{with_barcode} Barcodes, {with_area} Flächen, {with_nutzung} Nutzungen")

    return gebaeude, etage, room_data


def sort_key(room):
    nr = room['raumnr']
    m = re.match(r'(\d+)\.(\d+)([a-z]?)', nr)
    if m:
        return (0, int(m.group(1)), int(m.group(2)), m.group(3))
    return (1, 0, 0, nr)


def extract_liegenschaft(pdf_file):
    """Extrahiert den Liegenschaftsnamen aus dem Unterordner-Pfad unter plaene/."""
    parts = os.path.normpath(pdf_file).split(os.sep)
    if 'plaene' in parts:
        idx = parts.index('plaene')
        if idx + 1 < len(parts) - 1:  # Es gibt einen Unterordner zwischen plaene/ und der Datei
            return parts[idx + 1]
    return 'Standard'


def _write_liegenschaft_sheet(ws, pdf_entries):
    """Schreibt die Daten einer Liegenschaft in ein Worksheet. Gibt Raumanzahl zurück."""
    ws.append(['Gebäude', None, 'Etagen', None, 'Raumnr.', 'Fläche (m²)', 'Nutzung', 'Barcode'])
    total = 0
    first_etage_per_geb = {}

    for gebaeude, etage, rooms in pdf_entries:
        sorted_rooms = sorted(rooms.values(), key=sort_key)
        if not sorted_rooms:
            continue

        is_first_geb = gebaeude not in first_etage_per_geb
        first_etage_per_geb.setdefault(gebaeude, True)

        for i, room in enumerate(sorted_rooms):
            flaeche = float(room['flaeche']) if room['flaeche'] else None
            row = [
                gebaeude if i == 0 and is_first_geb else None,
                None,
                etage if i == 0 else None,
                None,
                room['raumnr'],
                flaeche,
                room['nutzung'],
                room['barcode']
            ]
            ws.append(row)
            total += 1

        if is_first_geb:
            first_etage_per_geb[gebaeude] = False

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 12
    return total


def write_xlsx(all_pdf_data, output_file, append=False):
    """Schreibt alle Raumdaten in eine xlsx — ein Sheet pro Liegenschaft.
    Bei append=True werden bestehende Sheets beibehalten, nur betroffene Liegenschaften ersetzt."""
    from collections import OrderedDict

    # Nach Liegenschaft gruppieren
    liegenschaften = OrderedDict()
    for gebaeude, etage, rooms, liegenschaft in all_pdf_data:
        if liegenschaft not in liegenschaften:
            liegenschaften[liegenschaft] = []
        liegenschaften[liegenschaft].append((gebaeude, etage, rooms))

    # Bestehende Datei laden oder neue erstellen
    if append and os.path.exists(output_file):
        wb = openpyxl.load_workbook(output_file)
        kept = [s for s in wb.sheetnames if s not in liegenschaften]
        print(f"  Append-Modus: {len(kept)} bestehende Sheets beibehalten, "
              f"{len(liegenschaften)} Liegenschaft(en) neu geschrieben")
        # Betroffene Sheets löschen (werden neu erstellt)
        for lieg_name in liegenschaften:
            if lieg_name[:31] in wb.sheetnames:
                del wb[lieg_name[:31]]
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    total_rooms = 0
    for lieg_name, pdf_entries in liegenschaften.items():
        ws = wb.create_sheet(title=lieg_name[:31])
        total_rooms += _write_liegenschaft_sheet(ws, pdf_entries)

    wb.save(output_file)
    total_sheets = len(wb.sheetnames)
    print(f"\n{output_file} geschrieben: {total_rooms} neue Räume, "
          f"{total_sheets} Sheet(s) gesamt, "
          f"Liegenschaften: {', '.join(wb.sheetnames)}")


def _format_duration(seconds):
    """Formatiert Dauer: '1h 23m' / '5m 12s' / '45s'."""
    h = int(seconds // 3600)
    m = int((seconds % 3600) // 60)
    s = int(seconds % 60)
    if h > 0:
        return f"{h}h {m:02d}m"
    if m >= 10:
        return f"{m}m"
    if m > 0:
        return f"{m}m {s:02d}s"
    return f"{s}s"


def limit_cpu(max_percent=90):
    """Begrenzt die CPU-Nutzung über Core-Affinity (z.B. 90% = 7 von 8 Cores)."""
    try:
        import psutil
        p = psutil.Process()
        all_cpus = list(range(psutil.cpu_count()))
        n_use = max(1, int(len(all_cpus) * max_percent / 100))
        p.cpu_affinity(all_cpus[:n_use])
        p.nice(psutil.BELOW_NORMAL_PRIORITY_CLASS if sys.platform == 'win32' else 10)
        print(f"  CPU begrenzt: {n_use}/{len(all_cpus)} Cores ({max_percent}%), Priorität gesenkt")
    except ImportError:
        print("  Hinweis: psutil nicht installiert, CPU-Limit nicht aktiv (pip install psutil)")
    except Exception as e:
        print(f"  Hinweis: CPU-Limit konnte nicht gesetzt werden: {e}")


def main():
    # Argumente parsen
    args = sys.argv[1:]
    output = OUTPUT_FILE
    pdf_files = []
    max_cpu = 90
    append = False

    i = 0
    while i < len(args):
        if args[i] == '-o' and i + 1 < len(args):
            output = args[i + 1]
            i += 2
        elif args[i] == '--cpu' and i + 1 < len(args):
            max_cpu = int(args[i + 1])
            i += 2
        elif args[i] == '--append':
            append = True
            i += 1
        else:
            pdf_files.append(args[i])
            i += 1

    # CPU begrenzen
    limit_cpu(max_cpu)

    # Wenn keine PDFs angegeben, alle *.pdf im Ordner /plaene (rekursiv) nehmen
    if not pdf_files:
        pdf_files = sorted(glob.glob('plaene/**/*.pdf', recursive=True))
        if not pdf_files:
            pdf_files = sorted(glob.glob('plaene/*.pdf'))
        if not pdf_files:
            pdf_files = sorted(glob.glob('*.pdf'))
        if not pdf_files:
            print("Keine PDF-Dateien gefunden!")
            print("Lege PDFs in den Ordner 'plaene/' oder gib sie als Argument an:")
            print("  python parse_gebaeudeplan.py plan1.pdf plan2.pdf ...")
            sys.exit(1)

    print(f"Verarbeite {len(pdf_files)} PDF(s): {', '.join(pdf_files)}")
    print(f"Ausgabe: {output}")

    # Kachelzahl aller PDFs vorab berechnen (schnell, nur Bildgröße lesen)
    import time
    print("\nBerechne Gesamtumfang...")
    total_tiles_all = 0
    for pdf_file in pdf_files:
        try:
            doc = fitz.open(pdf_file)
            page = doc[0]
            # Bildgröße bei DPI berechnen ohne tatsächlich zu rendern
            W = int(page.rect.width * DPI / 72)
            H = int(page.rect.height * DPI / 72)
            doc.close()
            total_tiles_all += count_tiles_for_image(W, H)
        except Exception:
            total_tiles_all += 48  # Fallback-Schätzung
    print(f"  {len(pdf_files)} PDFs, {total_tiles_all} Kacheln gesamt")

    # Globalen Fortschritt initialisieren
    _global_progress['total'] = total_tiles_all
    _global_progress['done'] = 0

    # OCR-Reader einmal initialisieren (spart Zeit bei mehreren PDFs)
    print("\nInitialisiere OCR-Reader...")
    reader = easyocr.Reader(['de', 'en'], gpu=False)

    # Alle PDFs verarbeiten
    all_pdf_data = []  # [(gebaeude, etage, rooms, liegenschaft), ...]
    t_total_start = time.time()
    _global_progress['t_start'] = t_total_start
    for pi, pdf_file in enumerate(pdf_files):
        t_pdf_start = time.time()
        if len(pdf_files) > 1:
            print(f"\n  >>> PDF {pi+1}/{len(pdf_files)} <<<")

        gebaeude, etage, rooms = process_single_pdf(pdf_file, reader)

        t_pdf_elapsed = time.time() - t_pdf_start
        print(f"  Dauer für diesen Plan: {_format_duration(t_pdf_elapsed)}")

        if rooms:
            liegenschaft = extract_liegenschaft(pdf_file)
            all_pdf_data.append((gebaeude, etage, rooms, liegenschaft))

    t_total_elapsed = time.time() - t_total_start

    if not all_pdf_data:
        print("\nKeine Raumdaten erkannt!")
        sys.exit(1)

    # Zusammenfassung
    total = sum(len(rooms) for _, _, rooms, _ in all_pdf_data)
    print(f"\n{'='*60}")
    print(f"GESAMT: {total} Räume aus {len(all_pdf_data)} PDFs")
    print(f"Gesamtdauer: {_format_duration(t_total_elapsed)}")
    print(f"{'='*60}")

    # Schreiben
    write_xlsx(all_pdf_data, output, append=append)

    # Aufräumen
    for f in ['_temp_plan_ocr.png', '_temp_tile.png']:
        if os.path.exists(f):
            os.remove(f)


if __name__ == '__main__':
    main()
    input("\nDrücke Enter zum Beenden...")
